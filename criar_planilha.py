import openpyxl

#criar planilha
book = openpyxl.Workbook()
'''
#visualizar paginas
print(book.sheetnames)


#como ciar uma pagina
book.create_sheet('Pagina1')
'''

#selecionar nome da pagina
nome_pagina = book['Sheet']


#adicionar dados na pagina
nome_pagina.append(['nome', 'quantidade', 'valor'])
nome_pagina.append(['teste1', '1', 'R$10,00'])
nome_pagina.append(['teste2', '2', 'R$20,00'])

#salvar
book.save('planilha.xlsx')

