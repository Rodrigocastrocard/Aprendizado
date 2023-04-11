import openpyxl

def ler():
    book = openpyxl.load_workbook('planilha.xlsx')
    pagina = book['nome']
    for linha in pagina.iter_rows(min_row=2):
        print(f'{linha[0].value},{linha[1].value},{linha[2].value}')
    """for cell in linha:
        print(cell.value)"""

def alterar():
    book = openpyxl.load_workbook('planilha.xlsx')
    pagina = book['nome']
    for linha in pagina.iter_rows(min_row=2):
        for cell in linha:
            if cell.value == 'teste1':
                cell.value = 'testealterado'
                print(cell.value)
                book.save('planilha2.xlsx')





